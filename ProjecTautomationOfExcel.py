import openpyxl
from openpyxl.chart import BarChart,Reference


wb = openpyxl.load_workbook('data.xlsx')
sheet = wb['Sheet1']      # sheet1 case sensitive
cell = sheet['a1']
cell1 = sheet.cell(1,1) #(row,col)
print(cell.value)
print(cell1.value)

print(sheet.max_column)
print(sheet.max_row)

for row in range(1, sheet.max_row+1):
    print(row)


for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row,3)
    print(cell.value)
    corrected_value = cell.value * 0.9
    corrected_value_cell =sheet.cell(row,4)
    corrected_value_cell.value = corrected_value

values = Reference(sheet, min_col = 4, min_row = 2,
                         max_col = 4, max_row = sheet.max_row)
chart = BarChart()
# Create object of BarChart class
chart = BarChart()
  
# adding data to the Bar chart object
chart.add_data(values)
  
# set the title of the chart
chart.title = " BAR-CHART "
  
# set the title of the x-axis
chart.x_axis.title = " X_AXIS "
  
# set the title of the y-axis
chart.y_axis.title = " Y_AXIS "

sheet.add_chart(chart,'a8')
wb.save('dataupdated.xlsx')